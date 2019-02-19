# Делаем необходимые ипорты
from vk_api import VkApi
from vk_api.upload import VkUpload
from vk_api.keyboard import VkKeyboard
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from sql import User
from time import sleep
from openpyxl import Workbook
from datetime import datetime
from traceback import format_exc
import os
import sys
import settings

# фикс бага с longpoll
def try_repeat(func):
    def wrapper(*args, **kwargs):
        while True:
            try:
                return func(*args, **kwargs)
            except:
                format_exc()
                sleep(15)
    return wrapper

def save_excel():
    wb = Workbook()
    ws = wb.active

    users = User.select().where(User.request_contact != '')

    ws['A1'] = "ID"
    ws['B1'] = "Имя"
    ws['C1'] = "Номер телефона"
    ws['D1'] = "Пол"
    ws['E1'] = "Контакты"
    ws['F1'] = "Наколдовал"
    ws['G1'] = "Запись на приём"

    for i, user in enumerate(users):
        ws['A' + str(2 + i)] = user.user_id
        ws['B' + str(2 + i)] = user.request_name
        ws['C' + str(2 + i)] = user.request_phone_number
        ws['D' + str(2 + i)] = user.request_gender
        ws['E' + str(2 + i)] = user.request_contact
        ws['F' + str(2 + i)] = user.request_meaning
        ws['G' + str(2 + i)] = user.request_record

    wb.save("upload.xlsx")



# Функция для получения информации о пользователе
def check_user(user_id, api):
    user = User.select().where(User.user_id == user_id).first()   
    if not user:
        user = User(user_id=user_id)
        user.save()
    user_info = api.users.get(user_ids=user.user_id, fields='city, bdate')[0]

    return user, user_info

# Функция для преобразования клавиатуры из строки в класс
def to_kb(kb):
    keyboard = VkKeyboard(one_time=True)
    kb = kb.split('\n')
    for i, k in enumerate(kb):
        kb[i] = k.split(';')
    if len(kb) == 1:
        for k in kb[0]:
            keyboard.add_button(k)
        return keyboard.get_keyboard()
    if len(kb) > 1:
        for k in kb:
            for p in k:
                keyboard.add_button(p)
            if k != kb[-1]:
                keyboard.add_line()
        return keyboard.get_keyboard()

# Функция проверки сообщения
def check_message(event, api, vk_session):
    # Получаем текст сообщения
    text = event.obj.text
    # Получаем данные пользователя
    user, user_info = check_user(event.obj.from_id, api)
    if user.user_id in settings.ADMINS.split(',') and text.lower() == "!анкета":
        save_excel()
        upload = VkUpload(vk_session)
        doc = upload.document_message(doc="upload.xlsx", title="upload.xlsx", peer_id=user.user_id)
        time = datetime.strftime(datetime.now(), '[%d.%m.%Y]')
        attachment = "doc" + str(doc[0]['owner_id']) + "_" + str(doc[0]['id'])
        api.messages.send(user_id=user.user_id, message=f"Анкеты{time}", attachment=attachment)
        return True
    elif user.user_id in settings.ADMINS.split(',') and text.lower() == "!рестарт":
        message = f"Перезапускаем бота..."
        api.messages.send(user_id=user.user_id, message=message)
        os.execl(sys.executable,*([sys.executable]+sys.argv))
    # Проверяем в каком отрезке сценария он находится
    elif user.level == "0,0,0":
        # Проверяем соответствует ли текст строке "начать"
        if text.lower() == "начать":
            # Указываем сообщение, которое будет отправлено
            message = f"🤖Приветствую {user_info['first_name']}, если помните меня - я тот самый бот Роберт."
            # Отправляем сообщение, user_id берем из базы данных, сообщение берем, которое указали ранее
            api.messages.send(user_id=user.user_id, message=message)
            # Тоже самое
            message = "Я ждал этого момента и обязательно проведу вас по нашему короткому демо, покажу функционал и подарю кое-что в конце 😊"
            api.messages.send(user_id=user.user_id, message=message)
            # Тоже самое
            message = "Напишите “Продолжить” и поехали! Не пишите кавычки."
            api.messages.send(user_id=user.user_id, message=message)
            # Переход к следующему участку сценария, меняем переменную level в базе данных, чтобы понимать, где находится пользователь
            user.level = "1,0,0"
            # Сохраняем внесенные изменения в базу данных
            user.save()
            # Возвращаем True, так как сообщение подошло
            return True
        else:
            message = "🤖Напишите “Начать” чтобы завести мой заржавелый механизм!"
            api.messages.send(user_id=user.user_id, message=message)
            return True
    elif user.level == "1,0,0":
        if text.lower() == "продолжить":
            message = "🤖Начнём с простого. 💌Вы сможете создавать рассылки с опросниками и отправлять их всем пользователям своей группы, чтобы напоминать о своих услугах и акциях, а также собирать данные клиентов."
            api.messages.send(user_id=user.user_id, message=message)
            # Тоже самое 
            message = "👉Нажмите на кнопку “Покажи пример!” или отправьте любое сообщение. в Демо используются кнопки ВКонтакте. Вы можете развернуть их, нажав на специальный значок расположенный в поле для ввода сообщения. 😉Вам стоит обновить мобильное приложение, если сейчас вы не видите этот значок."
            # Объявляем клавиатуру. указывается через функцию to_kb, в значении передается строка. Кнопки клавиатуры разделяются
            # знаком ;, если нужно, чтобы кнопка была с новой строки, ставится \n
            keyboard = to_kb("Покажи пример!")
            # Отправляем сообщение с клавиатурой
            api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

            user.level = "1,1,0"
            user.save()
            return True
    elif user.level == "1,1,0":
        message = "👆Псс… Ты ещё не забыл о нас? Компания “Думбльдор” дарит вам бесплатный набор волшебника✨. Только нам необходимо узнать вас получше перед тем как вы заберёте его"
        api.messages.send(user_id=user.user_id, message=message)
        message = "Ваш пол - напишите одно из слов\n1.🙇Мужской\n2.🙋Женский"
        api.messages.send(user_id=user.user_id, message=message)

        user.level = "1,1,1"
        user.save()
        return True
    elif user.level == "1,1,1":
        if text.lower() in ["мужской", "женский"]:
            user.request_gender = text.lower()
            message = "🙌Что бы вы наколдовали, если у вас была волшебная палочка?"
            api.messages.send(user_id=user.user_id, message=message)

            user.level = "1,1,2"
            user.save()
        else:
            message = "Хм, такой ответ я не приму. Выберите один вариант из двух."
            api.messages.send(user_id=user.user_id, message=message)
        return True
    elif user.level == "1,1,2":
        user.request_meaning = text

        message = f"Ваш ответ “{user.request_gender}”, “{user.request_meaning}” был записан, спасибо😊"
        
        api.messages.send(user_id=user.user_id, message=message)

        message = "🤖Нажмите на кнопку либо отправьте любое сообщение чтобы продолжить демо и взглянуть на пример Меню"
        keyboard = to_kb("Продолжить")
        api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

        user.level = "2,0,0"
        user.save()
        return True
    elif user.level == "2,0,0":
        message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
        keyboard = to_kb("О нас;Часто задаваемые вопросы\nПочему мы?;Скинуть гифку")
        api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

        user.level = "2,1,0"
        user.save()
        return True
    # Меню
    elif user.level == "2,1,0":
        if text == "О нас":
            message = "😜Мы молодые, энергичные ребята и обожаем работать на границе между людьми и программированием. ❤Мы любим своё дело, ведь боты - это же просто чудо!💥 Мы верим в то, что чат-боты смогут взять на себя очень многие рутинные процессы, улучшить взаимодействие с клиентами, решить огромное количество проблем и сделать вас счастливым😊"
            keyboard = to_kb("Меню;Продолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        elif text == "Часто задаваемые вопросы":
            message = "🤖Я объединил вопросы, которые не могут не интересовать вас!"
            keyboard = to_kb("Роберт, есть ли у тебя чувства?\nРоберт, где ты находишься?\nМеню\nПродолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,2"
            user.save()
            return True
        elif text == "Почему мы?":
            message = "💪А кто же ещё? - ха-ха, ладно… Мы совсем молодые ребята, но мы любим своё дело. Нас отличает от конкурентов то, что для нас - ваш проект будет центром вселенной. Мы будем с вами абсолютно открытыми, выйдем на связь в любое время, будем гибкими и креативными."
            keyboard = to_kb("Меню;Продолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        elif text == "Скинуть гифку":
            attachment = "doc454025337_486021163"
            keyboard = to_kb("Меню;Продолжить демо")
            api.messages.send(user_id=user.user_id, attachment=attachment, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        else:
            message = "Извините, но такого пункта нет в меню."
            keyboard = to_kb("О нас;Часто задаваемые вопросы\nПочему мы?;Скинуть гифку")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
    # О нас и Почему мы
    elif user.level == "2,1,1":
        if text == "Меню":
            message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
            keyboard = to_kb("О нас;Часто задаваемые вопросы\nПочему мы?;Скинуть гифку")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
        elif text == "Продолжить демо":
            message = "🤖Вы даже сможете записывать клиентов на приём!"
            keyboard = to_kb("Пример")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "3,0,0"
            user.save()
            return True
        else:
            message = "Хм, видимо вы ошиблись."
            keyboard = to_kb("Меню;Продолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
    # Часто задаваемые вопросы
    elif user.level == "2,1,2":
        if text == "Роберт, есть ли у тебя чувства?":
            message = "🤖эх, мои создатели не додумались до этого, но вы мне определенно нравитесь :)"
            keyboard = to_kb("Роберт, есть ли у тебя чувства?\nРоберт, где ты находишься?\nМеню\nПродолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
        elif text == "Роберт, какова твоя жизнь?":
            message = "🤖Ну знаете ли... Люблю общаться с людьми, работаю 24/7, живу на сервере. Не жизнь, а сказка"
            keyboard = to_kb("Роберт, есть ли у тебя чувства?\nРоберт, где ты находишься?\nМеню\nПродолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
        elif text == "Меню":
            message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
            keyboard = to_kb("О нас;Часто задаваемые вопросы\nПочему мы?;Скинуть гифку")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
        elif text == "Продолжить демо":
            message = "🤖Вы даже сможете записывать клиентов на приём!"
            keyboard = to_kb("Пример")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "3,0,0"
            user.save()
            return True
        else: 
            message = "Хм, видимо вы ошиблись."
            keyboard = to_kb("Роберт, есть ли у тебя чувства?\nРоберт, где ты находишься?\nМеню\nПродолжить демо")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,2"
            user.save()
            return True
    # Запись на приём
    elif user.level == "3,0,0":
        message = f"👋Ну здравствуй, {user_info['first_name']}! Айда к нам на приём, родненький"
        keyboard = to_kb("Записаться на прием")
        api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

        user.level = "3,1,0"
        user.save()
        return True

    elif user.level == "3,1,0":
        message = "📞Введите ваш номер телефона в формате +79181001010"
        api.messages.send(user_id=user.user_id, message=message)

        user.level = "3,1,1"
        user.save()
        return True
    elif user.level == "3,1,1":
        if text[0] == "+":
            text = text[1:]
        try:
            number = int(text)
            if len(text) == 11 and number > 0:
                message = "🌈Удобный день недели"
                keyboard = to_kb("Пн;Вт;Ср\nЧт;Пт")
                api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

                user.request_phone_number = text
                user.level = "3,1,2"
                user.save()
                return True
            else:
                message = "Номер телефона указан в неверном формате"
                print(text)
                api.messages.send(user_id=user.user_id, message=message)
                return True
        except ValueError:
            message = "Номер телефона указан в неверном формате"
            print(text)
            api.messages.send(user_id=user.user_id, message=message)
            return True

    elif user.level == "3,1,2":
        if text in ["Пн", "Вт", "Ср", "Чт", "Пт"]:
            message = "Последний шаг. Во сколько бы вы хотели придти (введите время в формате 00:00)"
            api.messages.send(user_id=user.user_id, message=message)

            user.request_record = text + ","
            user.level = "3,1,3"
            user.save()
            return True
        else:
            message = "Такого дня не было в списке, выберите один из вариантов."
            keyboard = to_kb("Пн;Вт;Ср\nЧт;Пт")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
    elif user.level == "3,1,3":
        if text[0] in ["0","1","2"]:
            try:
                hours = int(text[0:2])
                minutes = int(text[3:5])
                if text[2] == ":" and len(text) == 5 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                    user.request_record += str(hours) + ":" + str(minutes)
                    day = user.request_record.split(",")[0]
                    time = text
                    message = f"Отлично-с ! Ждём вас у нас в {day}, {time}. Мы вам напомним о приёме в сообщениях ВК😊"
                    api.messages.send(user_id=user.user_id, message=message)

                    message = "🤖Учтите что это всего лишь скромное Демо. Возможности чат-бота ограничены только вашей фантазией и функционалом мессенджера🌈"
                    keyboard = keyboard = to_kb("Познать мощь")
                    api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)
                    user.level = "3,1,4"
                    user.save()
                    return True
                else:
                    message = "Время указано в неверном формате."
                    api.messages.send(user_id=user.user_id, message=message)
                    return True
            except ValueError:
                message = "Время указано в неверном формате."
                api.messages.send(user_id=user.user_id, message=message)
                return True
            except IndexError:
                message = "Время указано в неверном формате."
                api.messages.send(user_id=user.user_id, message=message)
                return True
        else:
            try:
                hours = int("0" + text[0])
                minutes = int(text[2:4])
                if text[1] == ":" and len(text) == 4 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                    user.request_record += str(hours) + ":" + str(minutes)
                    day = user.request_record.split(",")[0]
                    time = text
                    message = f"Отлично-с ! Ждём вас у нас в {day}, {time}. Мы вам напомним о приёме в сообщениях ВК😊"
                    api.messages.send(user_id=user.user_id, message=message)

                    message = "🤖Учтите что это всего лишь скромное Демо. Возможности чат-бота ограничены только вашей фантазией и функционалом мессенджера🌈"
                    keyboard = keyboard = to_kb("Познать мощь")
                    api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)
                    user.level = "3,1,4"
                    user.save()
                    return True
                else:
                    message = "Время указано в неверном формате."
                    api.messages.send(user_id=user.user_id, message=message)
                    return True
            except ValueError:
                message = "Время указано в неверном формате."
                api.messages.send(user_id=user.user_id, message=message)
                return True
            except IndexError:
                message = "Время указано в неверном формате."
                api.messages.send(user_id=user.user_id, message=message)
                return True
            
    elif user.level == "3,1,4":
            message = "😇Возможный функционал:"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🤖чат-бот сможет создать сделку в вашей CRM системе"
            api.messages.send(user_id=user.user_id, message=message)

            message = "✅Сможет принять оплату"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💨Отгрузить все накопленные данные клиентов в файл Excel"
            api.messages.send(user_id=user.user_id, message=message)

            message = "👀Осуществлять поиск информации по вашему ресурсу"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💬Общаться со сторонними серверами с помощью API, предоставляя огромное количество всевозможного функционала. Например, бот переведёт сообщение на другой язык или покажет баланс на лицевом счёте клиента"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🔎Сумеет найти и сопоставить ключевые слова в длинном сообщении клиента и предоставить заготовленный ответ по этим ключевым словам"
            keyboard = to_kb("Познать больше мощи!")
            api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

            user.level = "3,1,5"
            user.save()
            return True

    elif user.level == "3,1,5":
            message = "🌈Вы сможете обновлять скрипты через самого чат-бота🔥"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🚀Вы сможете собирать статистику взаимодействия клиентов с вашим чат-ботом - в какие разделы клиенты заходят чаще всего, длительность диалога и так далее"
            api.messages.send(user_id=user.user_id, message=message)

            message = "Да и много чего интересного;)"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💪И конечно же мы пишем чат-ботов для telegram, facebook messenger, viber, и для веб-сайтов."
            keyboard = to_kb("Получить подарок!")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "4,0,0"
            user.save()
            return True
            

    elif user.level == "4,0,0":
            message = f"😅Слуушайте, {user_info['first_name']} ну вы молодчина! Мы предоставим первый месяц тех. поддержки вашего чат-бота бесплатно, на случай если нужно будет что-то подправить или изменить! 🌝Уточните свой номер телефона и лучший способ связаться с вами"
            api.messages.send(user_id=user.user_id, message=message)

            message = f"Это ваш номер телефона? - +{user.request_phone_number}"
            keyboard = to_kb("Да;Нет")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.request_name = user_info['first_name']
            user.level = "4,1,1"
            user.save()
            return True
    elif user.level == "4,1,1":
        if text == "Да":
            message = "😌Как к вам лучше будет достучаться? (Viber, WhatsApp, Telegram)"
            keyboard = to_kb("Viber;WhatsApp\nTelegram")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "4,2,0"
            user.save()
            return True
        elif text == "Нет":
            message = "Ваш номер телефона?"
            api.messages.send(user_id=user.user_id, message=message)

            user.level = "4,1,2"
            user.save()
            return True
    elif user.level == "4,1,2":
        if text[0] == "+":
            text = text[1:]
        try:
            number = int(text)
            if len(text) == 11:
                message = "😌Как к вам лучше будет достучаться? (Viber, WhatsApp, Telegram)"
                keyboard = to_kb("Viber;WhatsApp\nTelegram")
                api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

                user.level = "4,2,0"
                user.save()
                return True
            else:
                message = "Номер телефона указан в неверном формате"
                print(text)
                api.messages.send(user_id=user.user_id, message=message)
                return True
        except ValueError:
            message = "Номер телефона указан в неверном формате"
            print(text)
            api.messages.send(user_id=user.user_id, message=message)
            return True

    elif user.level == "4,2,0":
        message = '🔥Отлично! Мы с вами свяжемся совсем скоро и вместе обсудим возможности внедрения чат-бота. Напишите “Начать сначала", если хотите повторить;)'
        api.messages.send(user_id=user.user_id, message=message)

        user.request_contact = text
        user.level = "4,3,0"
        user.save()
        return True
    elif user.level == "4,3,0":
        if text.lower() == "начать сначала":
            message = "Вжух... Возвращаемся."
            keyboard = to_kb("Начать")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "0,0,0"
            user.save()

            return True
        
    return False

@try_repeat
def start():
    vk_session = VkApi(token=settings.BOT_TOKEN)
    api = vk_session.get_api()

    longpoll = VkBotLongPoll(vk_session, settings.GROUP_ID)

    for event in longpoll.listen():
        if event.type == VkBotEventType.MESSAGE_NEW:
            if check_message(event, api, vk_session) == False:
                message = "Хм, видимо вы ошиблись."
                api.messages.send(user_id=event.obj.from_id, message=message)




if __name__ == "__main__":
    start()
